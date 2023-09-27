import datetime
import os
import types

import openpyxl

import FSM
import config
from aiogram import Router, html, Bot, F
from aiogram.filters import CommandStart, Command
from aiogram.types import Message, CallbackQuery, InputFile, FSInputFile
from aiogram.fsm.context import FSMContext
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession
from markups import admin_markup as admin_markup
from database import User, Subscribe, Payments




admin_router = Router()


@admin_router.message(Command("admin"))
async def start_admin(message: Message):
    if message.from_user.id in config.ADMINS:
        await message.answer("Панель администратора", reply_markup=admin_markup.create_main_markup())


@admin_router.callback_query(F.data == "get_excel")
async def get_excel(callback: CallbackQuery, session: AsyncSession):
    workbook = openpyxl.Workbook()
    ws = workbook.active
    titles = ["ID Пользователя", "Ник пользователя", "ФИО", "Телефон", "Почта", "Статус подписки", "Дата оформления подписки", "Дата окончания подписки"]
    for i, value in enumerate(titles, 1):
        ws.cell(row=1, column=i).value = value
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 25
    result = await session.execute(select(User))
    users = result.fetchall()
    index = 2
    for user in users:
        user = user[0]
        subscribe_result = await session.execute(select(Subscribe).where(Subscribe.user_id == user.user_id and Subscribe.is_active == True))
        is_subscribe = subscribe_result.one_or_none()
        if is_subscribe is None:
            row = [user.user_id, user.username, user.fio, user.phone, user.email, "Нет подписки", "-", "-"]
        else:
            row = [user.user_id, user.username, user.fio, user.phone, user.email, "Есть подписка", is_subscribe[0].date_start, is_subscribe[0].date_end]
        for i, value in enumerate(row, 1):
            ws.cell(row=index, column=i).value = value
        index += 1
    path = f"Выгрузка {datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
    workbook.save(path)
    await callback.message.answer_document(document=FSInputFile(path))
    os.remove(path)


@admin_router.callback_query(F.data == "stats")
async def stats(callback: CallbackQuery, session: AsyncSession):

    result = await session.execute(select(User))
    count_users = result.fetchall()
    await callback.message.answer(f"Общее количество пользователей - {len(count_users)}")



